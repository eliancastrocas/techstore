from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from .models import Product, Service, FormRequest, StockMovement
from .forms import ProductForm, ServiceForm, FormRequestForm
from users.models import UserProfile
from django.db.models import Q
from django.http import HttpResponse
from django.conf import settings
import os
import csv
from .dian_api import get_cities, get_departments

# Create your views here.

from cart.models import Cart


def product_search(request):
    query = request.GET.get("q", "").strip()
    category = request.GET.get("category", "").strip()
    min_price = request.GET.get("min_price", "").strip()
    max_price = request.GET.get("max_price", "").strip()
    in_stock = request.GET.get("in_stock", "").strip()

    # Search only vendor's products if logged in as vendor
    if request.user.is_authenticated and hasattr(request.user, 'profile') and request.user.profile.is_vendedor():
        products = Product.objects.filter(seller=request.user.profile).order_by("-created_at")
    else:
        # Clients and guests see available products
        products = Product.objects.filter(stock__gt=0).order_by("-created_at")

    if query:
        products = products.filter(
            Q(name__icontains=query) | Q(description__icontains=query)
        )

    if category:
        products = products.filter(category=category)

    if min_price:
        try:
            products = products.filter(price__gte=float(min_price))
        except ValueError:
            pass

    if max_price:
        try:
            products = products.filter(price__lte=float(max_price))
        except ValueError:
            pass

    if in_stock:
        products = products.filter(stock__gt=0)

    cart = Cart.get_cart(request)
    return render(
        request,
        "products/product_list.html",
        {
            "products": products,
            "query": query,
            "category": category,
            "min_price": min_price,
            "max_price": max_price,
            "in_stock": in_stock,
            "cart": cart,
            "cart_count": cart.item_count,
        },
    )


def product_catalog(request):
    """Public product catalog for clients and guests"""
    query = request.GET.get("q", "").strip()
    category = request.GET.get("category", "").strip()
    min_price = request.GET.get("min_price", "").strip()
    max_price = request.GET.get("max_price", "").strip()

    products = Product.objects.filter(stock__gt=0).order_by("-created_at")

    if query:
        products = products.filter(
            Q(name__icontains=query) | Q(description__icontains=query)
        )

    if category:
        products = products.filter(category=category)

    if min_price:
        try:
            products = products.filter(price__gte=float(min_price))
        except ValueError:
            pass

    if max_price:
        try:
            products = products.filter(price__lte=float(max_price))
        except ValueError:
            pass

    cart = Cart.get_cart(request)
    return render(
        request,
        "products/product_catalog.html",
        {
            "products": products,
            "query": query,
            "category": category,
            "min_price": min_price,
            "max_price": max_price,
            "cart": cart,
            "cart_count": cart.item_count,
        },
    )


@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and u.profile.role == "vendedor")
def product_list(request):
    profile = request.user.profile
    products = Product.objects.filter(seller=profile).order_by("-created_at")

    # Get filter parameters from GET request
    query = request.GET.get("q", "").strip()
    category = request.GET.get("category", "").strip()
    min_price = request.GET.get("min_price", "").strip()
    max_price = request.GET.get("max_price", "").strip()
    in_stock = request.GET.get("in_stock", "").strip()

    if query:
        products = products.filter(
            Q(name__icontains=query) | Q(description__icontains=query)
        )

    if category:
        products = products.filter(category=category)

    if min_price:
        try:
            products = products.filter(price__gte=float(min_price))
        except ValueError:
            pass

    if max_price:
        try:
            products = products.filter(price__lte=float(max_price))
        except ValueError:
            pass

    if in_stock:
        products = products.filter(stock__gt=0)

    cart = Cart.get_cart(request)
    return render(request, "products/product_list.html", {
        "products": products,
        "query": query,
        "category": category,
        "min_price": min_price,
        "max_price": max_price,
        "in_stock": in_stock,
        "cart": cart,
        "cart_count": cart.item_count,
    })





from django.contrib.auth.decorators import user_passes_test


@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and u.profile.is_vendedor())
def product_create(request):
    try:
        profile = request.user.profile
    except UserProfile.DoesNotExist:
        profile = None

    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product_name = form.cleaned_data.get("name", "").strip()
            category = form.cleaned_data.get("category", "otro")
            price = form.cleaned_data.get("price", 0)
            description = form.cleaned_data.get("description", "")
            image_url = form.cleaned_data.get("image_url", "")
            stock = form.cleaned_data.get("stock", 0)
            is_damaged = form.cleaned_data.get("is_damaged", False)

            product, created = Product.objects.update_or_create(
                name=product_name,
                seller=profile,
                defaults={
                    "category": category,
                    "price": price,
                    "description": description,
                    "image_url": image_url,
                    "stock": stock,
                    "is_damaged": is_damaged,
                },
            )

            if created:
                # Only create movimento if there's NO existing entrada for this product
                if product.stock > 0 and profile:
                    StockMovement.objects.create(
                        product=product,
                        movement_type="entrada",
                        quantity=product.stock,
                        reason="Producto criado",
                        created_by=profile,
                    )
                messages.success(request, "Producto creado exitosamente.")
            else:
                messages.success(request, "Producto actualizado exitosamente.")
            return redirect("product_list")
    else:
        form = ProductForm()
    return render(
        request, "products/product_form.html", {"form": form, "action": "Criar"}
    )


@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and u.profile.is_vendedor())
def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    # Get user profile
    try:
        profile = request.user.profile
    except UserProfile.DoesNotExist:
        profile = None

    # Asegurar que el vendedor solo edite sus propios productos
    if profile and product.seller_id != profile.id:
        messages.error(request, "No tienes permisos para editar este producto.")
        return redirect("product_list")

    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            # guardar TODOS los campos (price, stock, etc.)
            form.save()
            messages.success(request, "Producto actualizado exitosamente.")
            return redirect("product_list")
    else:
        form = ProductForm(instance=product)

    return render(
        request,
        "products/product_form.html",
        {"form": form, "product": product, "action": "Editar"},
    )



@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and u.profile.is_vendedor())
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    # Get user profile
    try:
        profile = request.user.profile
    except UserProfile.DoesNotExist:
        profile = None

    if request.method == "POST":
        product.delete()
        messages.success(request, "Producto eliminado exitosamente.")
        return redirect("product_list")
    return render(request, "products/product_confirm_delete.html", {"product": product})


def services_catalog(request):
    """Página de servicios"""
    from cart.models import Cart
    from .models import Service

    cart = Cart.get_cart(request)
    services = Service.objects.all()
    return render(
        request,
        "products/services.html",
        {"cart": cart, "cart_count": cart.item_count, "services": services},
    )


@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and getattr(u.profile, "role", None) == "vendedor")
def vendor_form_requests(request):
    """Vendor dashboard for form requests and warranty claims"""
    from orders.models import WarrantyClaim

    profile = request.user.profile
    form_requests = FormRequest.objects.filter(status="pending").order_by("-created_at")
    warranty_claims = (
        WarrantyClaim.objects.filter(order__items__product__seller=profile)
        .distinct()
        .order_by("-created_at")
    )

    return render(
        request,
        "products/vendor_form_requests.html",
        {
            "form_requests": form_requests,
            "warranty_claims": warranty_claims,
            "object_list": form_requests,
        },
    )


@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and getattr(u.profile, "role", None) == "vendedor")
def bulk_upload(request):
    profile = request.user.profile
    new_products = []

    if request.method == "POST":
        file = request.FILES.get("file")
        if file:
            ext = file.name.split(".")[-1].lower()
            temp_path = os.path.join(settings.MEDIA_ROOT, "bulk", file.name)
            os.makedirs(os.path.dirname(temp_path), exist_ok=True)

            with open(temp_path, "wb") as f:
                for chunk in file.chunks():
                    f.write(chunk)

            rows = []
            if ext == "csv":
                try:
                    with open(temp_path, "r", encoding="utf-8") as f:
                        reader = csv.DictReader(f)
                        rows = list(reader)
                except Exception as e:
                    messages.error(request, f"Error: {str(e)}")
            elif ext in ["xlsx", "xls"]:
                try:
                    from openpyxl import load_workbook

                    wb = load_workbook(temp_path, data_only=True)
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        rows.append(dict(zip(headers, row)))
                except ImportError:
                    messages.error(request, "Usa solo CSV. Excel no disponible.")
                except Exception as e:
                    messages.error(request, f"Error Excel: {str(e)}")
            else:
                messages.error(request, "Usa CSV o Excel.")

            if not rows:
                messages.error(request, "El archivo está vacío.")
            else:
                required = ["name", "price", "description", "stock"]
                first = rows[0] if rows else {}
                missing = [c for c in required if c not in first]
                if missing:
                    messages.error(request, f"Columnas faltantes: {', '.join(missing)}")
                else:
                    for row in rows:
                        name = str(row.get("name", "")).strip()
                        price = float(row.get("price", 0) or 0)
                        description = str(row.get("description", "")).strip()
                        stock = int(row.get("stock", 0) or 0)
                        category = str(row.get("category", "otro")).strip().lower()
                        image_url = str(row.get("image_url", "")).strip()

                        if name and price > 0:
                            stock_qty = abs(stock)
                            product, created = Product.objects.update_or_create(
                                name=name,
                                seller=profile,
                                defaults={
                                    "price": abs(price),
                                    "description": description,
                                    "stock": stock_qty,
                                    "category": category if category else "otro",
                                    "image_url": image_url if image_url else "",
                                },
                            )
                            if stock_qty > 0:
                                StockMovement.objects.create(
                                    product=product,
                                    movement_type="entrada",
                                    quantity=stock_qty,
                                    reason="Carga masiva",
                                    created_by=profile,
                                )
                            if created:
                                new_products.append(name)
                            else:
                                new_products.append(f"{name} (actualizado)")

                    new_count = sum(
                        1 for p in new_products if not p.endswith("(actualizado)")
                    )
                    updated_count = len(new_products) - new_count
                    msg = f"{new_count} productos importados."
                    if updated_count > 0:
                        msg += f" {updated_count} actualizados."
                    messages.success(request, msg)

    return render(request, "products/bulk_upload.html", {"new_products": new_products})


def cities_list(request):
    """Lista de ciudades colombianas desde API externa (con búsqueda)"""
    query = request.GET.get("q", "").strip().lower()
    
    all_cities = get_cities()
    all_departments = get_departments()
    dept_map = {d["id"]: d["name"] for d in all_departments}
    
    # Add dept_name to each city
    for city in all_cities:
        city["dept_name"] = dept_map.get(city.get("departmentId"), "Desconocido")
    
    if query:
        cities = [city for city in all_cities if query in city["name"].lower() or query in city["dept_name"].lower()]
    else:
        cities = all_cities
    
    context = {
        "cities": cities,
        "query": request.GET.get("q", ""),
        "count": len(cities),
        "total": len(all_cities),
    }
    return render(request, "products/cities.html", context)


def reparacion_form(request):
    """Formulario específico para reparación de dispositivos"""
    from cart.models import Cart

    cart = Cart.get_cart(request)

    if request.method == "POST":
        form = FormRequestForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "¡Solicitud de reparación enviada exitosamente! Te contactaremos en 24 horas.")
            return redirect("services_catalog")
        else:
            messages.error(request, "Por favor corrija los errores en el formulario.")
    else:
        form = FormRequestForm()

    return render(request, "products/reparacion_form.html", {
        "form": form, 
        "cart": cart, 
        "cart_count": cart.item_count
    })


def mantenimiento_form(request):
    """Catálogo de mantenimiento (sin formulario). Agrega al carrito por GET con service_key."""
    from cart.models import Cart, CartItem
    from django.conf import settings
    from django.contrib.contenttypes.models import ContentType

    cart = Cart.get_cart(request)
    cart_count = cart.item_count

    def add_service_to_cart(service_key: str):
        if not cart:
            return

        # Mapa clave -> lista de posibles valores dentro de `Service.name`.
        # (El `name` en BD puede variar, por eso se usa lista + OR.)
        mapping = {
            "celular": ["celular", "telefono", "tel", "cel"],
            "laptop": ["laptop", "notebook", "portatil"],
            "tablet": ["tablet", "tab"],
            "audifonos": ["audif", "audífonos", "headphone", "earphone"],
            "revision_tecnica": ["revision", "revisi", "tecnica", "técnica"],
            "limpieza_profunda": ["limpieza", "profunda", "deep"],
        }

        candidates = mapping.get(service_key, ["mantenimiento"])

        # Construir búsqueda OR.
        q = Q()
        for c in candidates:
            q |= Q(name__icontains=c)

        service_qs = Service.objects.filter(q)
        if not service_qs.exists():
            service_qs = Service.objects.filter(name__icontains="mantenimiento")

        if not service_qs.exists():
            messages.warning(
                request,
                "No se encontró el servicio en el catálogo. Intenta con otro tipo.",
            )
            return

        service = service_qs.first()
        content_type = ContentType.objects.get_for_model(Service)

        cart_item, created = CartItem.objects.get_or_create(
            cart=cart,
            content_type=content_type,
            object_id=service.id,
            defaults={"quantity": 1},
        )
        if not created:
            cart_item.quantity += 1
            cart_item.save()

        messages.success(request, f"{service.name} agregado al carrito")


    # Botón "Agregar al carrito" (GET)
    if request.method == "GET" and request.GET.get("add_to_cart") == "on":
        service_key = (request.GET.get("service_key") or "").strip()
        add_service_to_cart(service_key)
        return redirect("services_catalog")

    # Mantener compatibilidad si alguien hace POST (no mostramos formulario en la plantilla)
    if request.method == "POST":
        messages.warning(request, "Este catálogo no tiene formulario. Usa el botón para agregar al carrito.")
        return redirect("services_catalog")

    # No renderizamos formulario real: enviamos placeholder
    form = FormRequestForm()
    return render(
        request,
        "products/mantenimiento_form.html",
        {"form": form, "cart": cart, "cart_count": cart_count},
    )



def warranty_request(request):
    """Formulario de garantía/reparación para clientes"""
    from cart.models import Cart

    cart = Cart.get_cart(request)

    if request.method == "POST":
        form = FormRequestForm(request.POST, request.FILES)
        if form.is_valid():
            form_request = form.save(commit=False)
            if request.user.is_authenticated:
                form_request.customer_name = (
                    request.user.get_full_name() or request.user.username
                )
                form_request.email = request.user.email
            form_request.save()
            messages.success(request, "¡Solicitud enviada! Te contactaremos pronto.")
            return redirect("home")
    else:
        form = FormRequestForm()

    return render(request, "products/warranty_form.html", {"form": form, "cart": cart})


@login_required
def warranty_request_list(request):
    """Lista de solicitudes de garantía del cliente"""
    form_requests = FormRequest.objects.filter(
        models.Q(seller=request.user.profile) | models.Q(status="pending")
    ).order_by("-created_at")

    return render(
        request, "products/warranty_list.html", {"form_requests": form_requests}
    )


@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and getattr(u.profile, "role", None) == "vendedor")
def service_create(request):
    profile = request.user.profile

    if request.method == "POST":
        form = ServiceForm(request.POST)
        if form.is_valid():
            service = form.save(commit=False)
            service.seller = profile
            service.save()
            messages.success(request, "Servicio creado exitosamente.")
            return redirect("service_list")
    else:
        form = ServiceForm()
    return render(request, "products/service_form.html", {"form": form, "action": "Crear"})


@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and getattr(u.profile, "role", None) == "vendedor")
def service_list(request):
    profile = request.user.profile
    services = Service.objects.filter(seller=profile).order_by("-created_at")
    return render(request, "products/service_list.html", {"services": services})


@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and getattr(u.profile, "role", None) == "vendedor")
def service_update(request, pk):
    service = get_object_or_404(Service, pk=pk, seller=request.user.profile)
    if request.method == "POST":
        form = ServiceForm(request.POST, instance=service)
        if form.is_valid():
            form.save()
            messages.success(request, "Servicio actualizado exitosamente.")
            return redirect("service_list")
    else:
        form = ServiceForm(instance=service)
    return render(request, "products/service_form.html", {"form": form, "action": "Editar", "service": service})


@login_required
@user_passes_test(lambda u: hasattr(u, "profile") and getattr(u.profile, "role", None) == "vendedor")
def service_delete(request, pk):
    service = get_object_or_404(Service, pk=pk, seller=request.user.profile)
    if request.method == "POST":
        service.delete()
        messages.success(request, "Servicio eliminado exitosamente.")
        return redirect("service_list")
    return render(request, "products/service_confirm_delete.html", {"service": service})
